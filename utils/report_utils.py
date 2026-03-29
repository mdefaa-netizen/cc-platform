import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

NAVY  = "1B2A4A"
TEAL  = "2A7F7F"
GOLD  = "C8963E"
CREAM = "F5F0E8"
WHITE = "FFFFFF"

# ── Excel ──────────────────────────────────────────────────────────────────────

def style_header_row(ws, row, col_count, fill_hex=NAVY):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=WHITE, size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def generate_excel(events, hosts, facilitators, feedback, date_label="All Events") -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Events Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Events Summary"
    ws1.append(["NH Humanities & CDFA - Community Conversations Program"])
    ws1["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws1.append([f"Report Period: {date_label}"])
    ws1.append([f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"])
    ws1.append([])

    headers = ["Event Name","Date","Time","City","Host","Venue","Facilitators",
               "Status","Attendance","Attendance Confirmed","Summary"]
    ws1.append(headers)
    style_header_row(ws1, ws1.max_row, len(headers))

    for e in events:
        facs = e.get("facilitator_names", "")
        ws1.append([
            e.get("event_name",""), e.get("event_date",""), e.get("event_time",""),
            e.get("city",""), e.get("host_name",""), e.get("venue_name",""),
            facs, e.get("status",""),
            e.get("attendance_count",""), "Yes" if e.get("attendance_confirmed") else "No",
            e.get("event_summary",""),
        ])
    auto_width(ws1)

    # ── Sheet 2: Attendance ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Attendance Data")
    ws2.append(["Attendance Summary"])
    ws2["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws2.append([])
    headers2 = ["Event Name","Date","City","Attendance Count","Confirmed"]
    ws2.append(headers2)
    style_header_row(ws2, ws2.max_row, len(headers2))
    total = 0
    for e in events:
        cnt = e.get("attendance_count") or 0
        total += cnt
        ws2.append([e.get("event_name",""), e.get("event_date",""), e.get("city",""),
                    cnt, "Yes" if e.get("attendance_confirmed") else "No"])
    ws2.append([])
    ws2.append(["TOTAL", "", "", total, ""])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True, color=NAVY)
    ws2.cell(ws2.max_row, 4).font = Font(bold=True, color=NAVY)
    auto_width(ws2)

    # ── Sheet 3: Feedback ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Feedback")
    ws3.append(["Participant Feedback"])
    ws3["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws3.append([])
    headers3 = ["Event","Date Submitted","Participant","Rating","Feedback"]
    ws3.append(headers3)
    style_header_row(ws3, ws3.max_row, len(headers3))
    for fb in feedback:
        ws3.append([fb.get("event_name",""), str(fb.get("submitted_date",""))[:10] if fb.get("submitted_date") else "",
                    fb.get("participant_name",""), fb.get("rating",""), fb.get("feedback_text","")])
    auto_width(ws3)

    # ── Sheet 4: Payments ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Payments")
    ws4.append(["Payment Summary"])
    ws4["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws4.append([])
    headers4 = ["Payee Name","Type","Check Payable To","Amount","Status","Payment Date","Notes"]
    ws4.append(headers4)
    style_header_row(ws4, ws4.max_row, len(headers4))
    total_paid = 0
    for h in hosts:
        ws4.append([h.get("name",""), "Host", h.get("check_payable_to",""),
                    f"${h.get('payment_amount',0):.2f}", h.get("payment_status",""),
                    h.get("payment_date","") or "", h.get("notes","") or ""])
        if h.get("payment_status") == "Paid":
            total_paid += h.get("payment_amount", 0)
    for f in facilitators:
        ws4.append([f.get("name",""), "Facilitator", f.get("check_payable_to",""),
                    f"${f.get('payment_amount',0):.2f}", f.get("payment_status",""),
                    f.get("payment_date","") or "", f.get("notes","") or ""])
        if f.get("payment_status") == "Paid":
            total_paid += f.get("payment_amount", 0)
    ws4.append([])
    ws4.append(["TOTAL PAID", "", "", f"${total_paid:.2f}", "", "", ""])
    ws4.cell(ws4.max_row, 1).font = Font(bold=True, color=NAVY)
    ws4.cell(ws4.max_row, 4).font = Font(bold=True, color=NAVY)
    auto_width(ws4)

    # ── Sheet 5: Contacts ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Contacts")
    ws5.append(["Contacts - Hosts & Facilitators"])
    ws5["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws5.append([])
    ws5.append(["— HOSTS —"])
    ws5[f"A{ws5.max_row}"].font = Font(bold=True, color=TEAL)
    headers5h = ["Name","Venue","Address","City","State","Zip","Contact Person","Email","Phone"]
    ws5.append(headers5h)
    style_header_row(ws5, ws5.max_row, len(headers5h), TEAL)
    for h in hosts:
        ws5.append([h.get("name",""), h.get("venue_name",""), h.get("address",""),
                    h.get("city",""), h.get("state","NH"), h.get("zip_code",""),
                    h.get("contact_person",""), h.get("email",""), h.get("phone","")])
    ws5.append([])
    ws5.append(["— FACILITATORS —"])
    ws5[f"A{ws5.max_row}"].font = Font(bold=True, color=TEAL)
    headers5f = ["Name","Email","Phone","Specialization","Notes"]
    ws5.append(headers5f)
    style_header_row(ws5, ws5.max_row, len(headers5f), TEAL)
    for f in facilitators:
        ws5.append([f.get("name",""), f.get("email",""), f.get("phone",""),
                    f.get("specialization","") or "", f.get("notes","") or ""])
    auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────────────

class ReportPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(27, 42, 74)
        self.cell(0, 10, "NH Humanities & CDFA Community Conversations", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, "Program Coordination Report", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(200, 150, 62)
        self.set_line_width(1)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()} | Generated {datetime.now().strftime('%B %d, %Y')}", align="C")

    def section_title(self, title):
        self.ln(4)
        self.set_fill_color(27, 42, 74)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def table_header(self, cols, widths):
        self.set_fill_color(42, 127, 127)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 8)
        for col, w in zip(cols, widths):
            self.cell(w, 7, col, border=1, fill=True)
        self.ln()
        self.set_text_color(0, 0, 0)
        self.set_font("Helvetica", size=8)

    def safe(self, text):
        """Strip characters outside latin-1 range."""
        return str(text).replace("-","-").replace("–","-").replace("\u2019","'").replace("\u201c",'"').replace("\u201d",'"').encode("latin-1","replace").decode("latin-1")

    def table_row(self, vals, widths, fill=False):
        self.set_fill_color(245, 240, 232)
        self.set_font("Helvetica", size=8)
        for val, w in zip(vals, widths):
            self.cell(w, 6, self.safe(str(val))[:40], border=1, fill=fill)
        self.ln()

def generate_pdf(events, hosts, facilitators, feedback, date_label="All Events") -> bytes:
    pdf = ReportPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Meta
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 6, f"Report Period: {date_label}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Events Summary
    pdf.section_title("EVENTS SUMMARY")
    cols   = ["Event Name","Date","City","Host","Status","Attendance"]
    widths = [50, 22, 28, 35, 22, 22]
    pdf.table_header(cols, widths)
    for i, e in enumerate(events):
        pdf.table_row([
            e.get("event_name","")[:30], e.get("event_date",""), e.get("city",""),
            e.get("host_name","")[:20], e.get("status",""), str(e.get("attendance_count","") or "-")
        ], widths, fill=(i % 2 == 0))

    # Attendance
    pdf.section_title("ATTENDANCE DATA")
    total_att = sum(e.get("attendance_count") or 0 for e in events)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, f"Total Attendance Across All Events: {total_att}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", size=9)
    for e in events:
        conf = "Confirmed" if e.get("attendance_confirmed") else "Unconfirmed"
        pdf.cell(0, 6, f"  - {e.get('event_name','')}: {e.get('attendance_count') or '-'} attendees - {conf}",
                 new_x="LMARGIN", new_y="NEXT")

    # Event Narratives
    narratives = [e for e in events if e.get("event_summary")]
    if narratives:
        pdf.section_title("EVENT SUMMARIES / NARRATIVES")
        for e in narratives:
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, f"{e.get('event_name','')} - {e.get('event_date','')}",
                     new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", size=8)
            pdf.multi_cell(0, 5, e.get("event_summary",""), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

    # Payments
    pdf.section_title("PAYMENT SUMMARY")
    total_paid  = sum((h.get("payment_amount",0) or 0) for h in hosts  if h.get("payment_status")=="Paid")
    total_paid += sum((f.get("payment_amount",0) or 0) for f in facilitators if f.get("payment_status")=="Paid")
    total_pend  = sum((h.get("payment_amount",0) or 0) for h in hosts  if h.get("payment_status")=="Pending")
    total_pend += sum((f.get("payment_amount",0) or 0) for f in facilitators if f.get("payment_status")=="Pending")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, f"Total Paid: ${total_paid:,.2f}   |   Total Pending: ${total_pend:,.2f}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    cols2   = ["Payee","Type","Amount","Status","Date Paid"]
    widths2 = [55, 25, 25, 25, 30]
    pdf.table_header(cols2, widths2)
    for i, h in enumerate(hosts):
        pdf.table_row([h.get("name",""), "Host", f"${h.get('payment_amount',0):.2f}",
                       h.get("payment_status",""), h.get("payment_date","") or "-"],
                      widths2, fill=(i%2==0))
    offset = len(hosts)
    for i, f in enumerate(facilitators):
        pdf.table_row([f.get("name",""), "Facilitator", f"${f.get('payment_amount',0):.2f}",
                       f.get("payment_status",""), f.get("payment_date","") or "-"],
                      widths2, fill=((i+offset)%2==0))

    # Feedback
    if feedback:
        pdf.section_title("PARTICIPANT FEEDBACK")
        for fb in feedback:
            pdf.set_font("Helvetica", "B", 8)
            rating_str = f"Rating: {fb.get('rating','N/A')}/5" if fb.get("rating") else ""
            pdf.cell(0, 6, f"{fb.get('event_name','')} - {fb.get('participant_name','Anonymous')} {rating_str}",
                     new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", size=8)
            pdf.multi_cell(0, 5, fb.get("feedback_text",""), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    # Contacts
    pdf.add_page()
    pdf.section_title("HOST CONTACTS")
    cols3   = ["Name","Venue","City","Contact Person","Email","Phone"]
    widths3 = [35, 35, 22, 30, 45, 22]
    pdf.table_header(cols3, widths3)
    for i, h in enumerate(hosts):
        pdf.table_row([h.get("name",""), h.get("venue_name",""), h.get("city",""),
                       h.get("contact_person",""), h.get("email",""), h.get("phone","")],
                      widths3, fill=(i%2==0))

    pdf.ln(4)
    pdf.section_title("FACILITATOR CONTACTS")
    cols4   = ["Name","Email","Phone","Specialization"]
    widths4 = [45, 60, 30, 55]
    pdf.table_header(cols4, widths4)
    for i, f in enumerate(facilitators):
        pdf.table_row([f.get("name",""), f.get("email",""), f.get("phone",""),
                       f.get("specialization","") or "-"],
                      widths4, fill=(i%2==0))

    return bytes(pdf.output())
